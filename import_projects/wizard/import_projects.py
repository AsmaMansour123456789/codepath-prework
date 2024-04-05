from odoo import models, fields, api
from xlrd import open_workbook, xldate_as_datetime
from odoo.exceptions import UserError
import re
from openpyxl import load_workbook
from base64 import b64decode
import base64
from datetime import datetime


class ProjectImportWizard(models.TransientModel):
    _name = 'project.import.wizard'
    _description = 'Import Projects from Excel'

    name = fields.Char(string='Name')
    file = fields.Binary(string='File')

    def import_projects(self):
        wb = open_workbook(file_contents=b64decode(self.file) or b'')
        sheet = wb.sheets()[0]

        for i in range(sheet.nrows):
            if i == 0:
                continue

            try:
                # Read data from Excel
                # id = int(sheet.cell(i, 0).value)
                sequence = sheet.cell(i, 1).value
                name = sheet.cell(i, 2).value
                worksheet_template_name = sheet.cell(i, 3).value
                user_id_name = sheet.cell(i, 4).value
                partner_id_name = sheet.cell(i, 5).value
                label_tasks = sheet.cell(i, 6).value
                company_name = sheet.cell(i, 7).value
                etiquettes = sheet.cell(i, 8).value.split(',')
                description = sheet.cell(i, 9).value
                user_id_name = sheet.cell(i, 10).value
                partner_phone = sheet.cell(i, 11).value
                partner_email = sheet.cell(i, 12).value
                analytic_account_id_name = sheet.cell(i, 13).value
                allow_timesheets = sheet.cell(i, 15).value
                allow_quotations = sheet.cell(i, 16).value
                rating_active = sheet.cell(i, 17).value
                allow_billable = sheet.cell(i, 18).value
                allow_material = sheet.cell(i, 19).value
                allow_forecast = sheet.cell(i, 20).value
                allow_subtasks = sheet.cell(i, 21).value
                allow_recurring_tasks = sheet.cell(i, 22).value
                services_sur_site = sheet.cell(i, 23).value

                # Find or create related records
                worksheet_template = self.env['worksheet.template'].search(
                    [('name', '=', worksheet_template_name)], limit=1)
                user = self.env['res.users'].search([('name', '=', user_id_name)], limit=1)
                partner = self.env['res.partner'].search([('name', '=', partner_id_name)], limit=1)
                company = self.env['res.company'].search([('name', '=', company_name)], limit=1)
                analytic_account = self.env['account.analytic.account'].search(
                    [('name', '=', analytic_account_id_name)], limit=1)

                # Find or create tags and link them to the project
                tag_ids = []
                for etiquette_name in etiquettes:
                    tag = self.env['project.tags'].search([('name', '=', etiquette_name.strip())], limit=1)
                    if not tag:
                        tag = self.env['project.tags'].create({'name': etiquette_name.strip()})
                    tag_ids.append(tag.id)

                if 'AGPM DOMAINE 2020' == name:
                    print(name)

                # Create project record
                project_vals = {
                    # 'id': id,
                    'sequence': sequence,
                    'name': name,
                    'worksheet_template_id': worksheet_template.id,
                    'user_id': user.id,
                    'partner_id': partner.id,
                    'label_tasks': label_tasks,
                    'company_id': company.id,
                    'tag_ids': [(6, 0, tag_ids)],
                    'description': re.sub('<.*?>', '', description),
                    'partner_phone': partner_phone,
                    'partner_email': partner_email,
                    'analytic_account_id': analytic_account.id,
                    'allow_timesheets': allow_timesheets,
                    'allow_quotations': allow_quotations,
                    'rating_active': rating_active,
                    'allow_billable': allow_billable,
                    'allow_material': allow_billable,
                    'allow_forecast': allow_forecast,
                    'allow_subtasks': allow_subtasks,
                    'allow_recurring_tasks': allow_recurring_tasks,
                    'is_fsm': services_sur_site,
                }
                try:
                    project = self.env['project.project'].create(project_vals)
                    self.env.cr.commit()
                except:
                    pass

            except Exception as e:
                # Handle exceptions or skip rows with errors
                pass

        return True

    # def import_projects(self):
    #     # wb = open_workbook(file_contents=self.file)
    #     wb = open_workbook(file_contents=b64decode(self.file) or b'')
    #     sheet = wb.sheets()[0]
    #
    #
    #     for i in range(sheet.nrows):
    #         if i == 0:
    #             continue
    #
    #         try:
    #             # Read data from Excel
    #             sequence = int(sheet.cell(i, 0).value)
    #             name = sheet.cell(i, 1).value
    #             worksheet_template_name = sheet.cell(i, 2).value
    #             user_name = sheet.cell(i, 3).value
    #             partner_name = sheet.cell(i, 4).value
    #             privacy_visibility = sheet.cell(i, 5).value
    #             subtask_project_name = sheet.cell(i, 6).value
    #             label_tasks = sheet.cell(i, 7).value
    #             company_name = sheet.cell(i, 8).value
    #             etiquette = sheet.cell(i, 9).value
    #             description = sheet.cell(i, 10).value
    #             user_name = sheet.cell(i, 11).value
    #             partner_phone = sheet.cell(i, 12).value
    #             partner_email = sheet.cell(i, 13).value
    #             analytic_account_name = sheet.cell(i, 14).value
    #             allow_timesheets = sheet.cell(i, 16).value
    #             allow_timesheet_timer = sheet.cell(i, 17).value
    #             allow_quotations = sheet.cell(i, 18).value
    #             rating_active = sheet.cell(i, 19).value
    #             allow_billable = sheet.cell(i, 20).value
    #             allow_material = sheet.cell(i, 21).value
    #             allow_forecast = sheet.cell(i, 22).value
    #             allow_subtasks = sheet.cell(i, 23).value
    #             allow_recurring_tasks = sheet.cell(i, 24).value
    #             allowed_portal_user_names = sheet.cell(i, 25).value
    #
    #             # Find or create related records
    #             worksheet_template = self.env['worksheet.template'].search(
    #                 [('name', '=', worksheet_template_name)], limit=1)
    #             user = self.env['res.users'].search([('name', '=', user_name)], limit=1)
    #             partner = self.env['res.partner'].search([('name', '=', partner_name)], limit=1)
    #             subtask_project = self.env['project.project'].search([('name', '=', subtask_project_name)], limit=1)
    #             company = self.env['res.company'].search([('name', '=', company_name)], limit=1)
    #             analytic_account = self.env['account.analytic.account'].search(
    #                 [('name', '=', analytic_account_name)], limit=1)
    #             # allowed_portal_users = self.env['res.users'].search(
    #             #     [('name', 'in', allowed_portal_user_names.split(','))])
    #
    #             etiquettes = etiquette.split(',')
    #
    #             # Find or create tags and link them to the project
    #             tag_ids = []
    #             for etiquette_name in etiquettes:
    #                 tag = self.env['project.tags'].search([('name', '=', etiquette_name.strip())], limit=1)
    #                 if not tag:
    #                     tag = self.env['project.tags'].create({'name': etiquette_name.strip()})
    #                 tag_ids.append(tag.id)
    #
    #             # Create project record
    #             project_vals = {
    #                 'sequence': sequence,
    #                 'name': name,
    #                 'worksheet_template_id': worksheet_template.id,
    #                 'user_id': user.id,
    #                 'partner_id': partner.id,
    #                 'privacy_visibility': 'portal',
    #                 # 'subtask_project_id': subtask_project.id,
    #                 'label_tasks': label_tasks,
    #                 'company_id': company.id,
    #                 'tag_ids': [(6, 0, tag_ids)],
    #                 'description': re.sub('<.*?>', '', description),
    #                 'partner_phone': partner_phone,
    #                 'partner_email': partner_email,
    #                 'analytic_account_id': analytic_account.id,
    #                 'allow_timesheets': allow_timesheets,
    #                 # 'allow_timesheet_timer': allow_timesheet_timer,
    #                 'allow_quotations': allow_quotations,
    #                 'rating_active': rating_active,
    #                 'allow_billable': allow_billable,
    #                 'allow_material': allow_material,
    #                 'allow_forecast': allow_forecast,
    #                 'allow_subtasks': allow_subtasks,
    #                 'allow_recurring_tasks': allow_recurring_tasks,
    #                 # 'allowed_portal_user_ids': [(6, 0, allowed_portal_users.ids)],
    #             }
    #             try:
    #
    #                 project = self.env['project.project'].create(project_vals)
    #                 self.env.cr.commit()
    #             except:
    #                 pass
    #
    #         except Exception as e:
    #             # Handle exceptions or skip rows with errors
    #             pass
    #
    #     return True
