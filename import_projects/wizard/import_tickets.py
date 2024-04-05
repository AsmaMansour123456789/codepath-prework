from odoo import models, fields, api
from xlrd import open_workbook, xldate_as_datetime
from odoo.exceptions import UserError
import re
from openpyxl import load_workbook
from base64 import b64decode
import base64
from datetime import datetime


class TicketImportWizard(models.TransientModel):
    _name = 'ticket.import.wizard'
    _description = 'Import Tickets from Excel'

    name = fields.Char(string='Name')
    file = fields.Binary(string='File')

    def import_helpdesk_tickets(self):
        wb = open_workbook(file_contents=base64.decodebytes(self.file))
        sheet = wb.sheets()[0]

        for i in range(sheet.nrows):
            if i == 0:
                continue

            try:
                display_name = sheet.cell(i, 0).value
                client = sheet.cell(i, 1).value
                assigned_to = sheet.cell(i, 2).value
                ticket_type = sheet.cell(i, 3).value
                priority = sheet.cell(i, 4).value
                labels = sheet.cell(i, 5).value
                activities = sheet.cell(i, 6).value
                stage = sheet.cell(i, 7).value
                company = sheet.cell(i, 8).value
                sla_status = sheet.cell(i, 9).value
                sla_status_display_name = sheet.cell(i, 10).value
                sla_created_date = sheet.cell(i, 11).value
                ticket_type_display_name = sheet.cell(i, 12).value
                label_name = sheet.cell(i, 13).value
                label_color = sheet.cell(i, 14).value
                project = sheet.cell(i, 15).value
                project_display_name = sheet.cell(i, 16).value
                task_display_name = sheet.cell(i, 17).value
                ticket_creation_date = sheet.cell(i, 18).value
                due_date = sheet.cell(i, 19).value
                client_email = sheet.cell(i, 20).value
                cc_email = sheet.cell(i, 21).value
                portal_access_url = sheet.cell(i, 22).value
                is_closed_by_partner = sheet.cell(i, 23).value
                sales_order_item_display_name = sheet.cell(i, 24).value
                article_name = sheet.cell(i, 25).value
                description = sheet.cell(i, 26).value
                time_sheets = sheet.cell(i, 27).value
                time_sheet_date = sheet.cell(i, 28).value
                time_sheet_employee_name = sheet.cell(i, 29).value
                time_sheet_employee_display_name = sheet.cell(i, 30).value
                time_sheet_timer = sheet.cell(i, 31).value
                time_sheet_description = sheet.cell(i, 32).value
                total_hours_passed = sheet.cell(i, 33).value

                # Convert date values
                if ticket_creation_date:
                    ticket_creation_date = xldate_as_datetime(ticket_creation_date, wb.datemode).date()
                else:
                    ticket_creation_date = False
                if due_date:
                    due_date = xldate_as_datetime(due_date, wb.datemode).date()
                if time_sheet_date:
                    time_sheet_date = xldate_as_datetime(time_sheet_date, wb.datemode).date()

                # Handle mappings and conditions if needed
                # Example: Map 'assigned_to' based on certain conditions
                if assigned_to == "Stephane Verbrugghe":
                    assigned_to_name = "Florence Pau"
                elif assigned_to == "Maxime longuet":
                    assigned_to_name = "Florence Pau"
                elif assigned_to == "Fabrice et Ludovic":
                    assigned_to_name = "Eric Laborde"
                elif assigned_to == "Victor Lima":
                    assigned_to_name = "Victor"
                elif assigned_to == "Services Clients et Fournisseurs":
                    assigned_to_name = "Service Administratif"
                else:
                    assigned_to_name = assigned_to

                tags = []
                tags_ids = []
                if company == "Pau Digital Factory":
                    tags.extend(["ERP", "WEB"])

                if company == 'Action Informatique':
                    tags.append("INFORMATIQUE")

                for tag in tags:
                    tag = self.env['helpdesk.tag'].search([('name', '=', tag)], limit=1)
                    tags_ids.append(tag.id)

                # Find or create related records (as before)
                assigned_user = self.env['res.users'].search([('name', '=', assigned_to_name)], limit=1)
                client_partner = self.env['res.partner'].search([('name', '=', client)], limit=1)
                no_project = False
                project_record = self.env['project.project'].search([('name', 'ilike', project)], limit=1)
                if not project_record:
                    project_record = self.env['project.project'].create({'name':project})
                    no_project = True
                    self.env.cr.commit()
                task_record = self.env['project.task'].search([('name', '=', task_display_name), ('project_id', '=', project_record.id)])
                ticket_type = self.env['helpdesk.ticket.type'].search([('name', '=', ticket_type_display_name)], limit=1)
                stage_id = self.env['helpdesk.stage'].search([('name', '=', stage)], limit=1)
                if priority == 'Priorité basse':
                    priority_id = '0'
                elif priority == 'Priorité moyenne':
                    priority_id = '1'
                elif priority == 'Haute priorité':
                    priority_id = '2'
                elif priority == 'Urgent':
                    priority_id = '3'
                else:
                    priority_id = False
                    # Add more search queries for related records as needed
                task_records = False
                if task_record:
                    task_records = [task_record.id]

                # Create or update the helpdesk ticket
                if display_name:
                    ticket = self.env['helpdesk.ticket'].search([('name', '=', display_name)], limit=1)
                    if no_project:
                        ticket.project_id = project_record.id
                    if not ticket:
                        ticket_vals = {
                            'name': display_name,
                            'partner_id': client_partner.id,
                            'user_id': assigned_user.id,
                            'ticket_type_id': ticket_type.id,
                            'priority': priority_id,
                            'stage_id': stage_id.id,
                            'company_id': self.env.company.id,
                            'tag_ids': tags_ids,
                            # 'sla_status': sla_status_display_name,
                            # 'sla_created_date': sla_created_date,
                            'description': description,
                            'assign_date': ticket_creation_date,
                            # 'date_deadline': due_date,
                            'partner_email': client_email,
                            'email_cc': cc_email,
                            'access_url': portal_access_url,
                            'closed_by_partner': is_closed_by_partner,
                            'project_id': project_record.id,
                            # 'fsm_task_ids': task_record.ids,
                            'fsm_task_ids': [(6, 0, task_record.id)],
                            # Add other fields for the ticket as needed
                        }


                        try:
                            ticket = self.env['helpdesk.ticket'].create(ticket_vals)
                            print(ticket)
                            ticket.write({'project_id': project_record.id, 'fsm_task_ids': [(6, 0, task_record.id)]})
                            print(ticket)
                            self.env.cr.commit()
                        except Exception as e:
                            # Handle exceptions or log errors
                            print(f"Error creating ticket: {str(e)}")
                else:
                    print(ticket)

                # Create time sheets if applicable
                if ticket and time_sheet_description:
                    if not ticket.task_ticket_id:
                        ticket.task_ticket_id = task_record.id
                    if not ticket.project_id:
                        ticket.project_id = project_record.id
                    existing_timesheet = self.env['account.analytic.line'].search([
                        ('helpdesk_ticket_id', '=', ticket.id),
                        ('name', '=', time_sheet_description),
                    ], limit=1)

                    if not existing_timesheet:

                        employee_id = self.env['hr.employee'].search([('name', '=', time_sheet_employee_display_name)],
                                                                     limit=1)
                        timesheet_vals = {
                            'helpdesk_ticket_id': ticket.id,
                            'name': time_sheet_description,
                            'date': time_sheet_date,
                            'account_id': ticket.project_id.analytic_account_id.id,
                            'unit_amount': total_hours_passed,
                            'employee_id': employee_id.id,
                            # Add other fields for the timesheet as needed
                        }

                        try:
                            timesheet = self.env['account.analytic.line'].create(timesheet_vals)
                            # ticket.write({'timesheet_ids': [(0, 0, timesheet_vals)]})
                            self.env.cr.commit()
                        except Exception as e:
                            # Handle exceptions or log errors
                            print(f"Error creating timesheet: {str(e)}")
                elif not display_name:
                    print(time_sheet_description)

            except Exception as e:
                # Handle exceptions or skip rows with errors
                print(f"Error processing row {i}: {str(e)}")

        return True
