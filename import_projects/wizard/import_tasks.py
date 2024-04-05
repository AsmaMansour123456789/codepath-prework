from odoo import models, fields, api
from xlrd import open_workbook, xldate_as_datetime
from odoo.exceptions import UserError
import re
from openpyxl import load_workbook
from base64 import b64decode
import base64
from datetime import datetime


class TaskImportWizard(models.TransientModel):
    _name = 'task.import.wizard'
    _description = 'Import Tasks from Excel'

    name = fields.Char(string='Name')
    file = fields.Binary(string='File')

    def import_tasks(self):
        wb = open_workbook(file_contents=base64.decodebytes(self.file))
        sheet = wb.sheets()[0]

        for i in range(sheet.nrows):
            if i == 0:
                continue

            try:
                title = sheet.cell(i, 0).value

                project = sheet.cell(i, 1).value
                assigned_to = sheet.cell(i, 2).value
                company = sheet.cell(i, 3).value
                activities = sheet.cell(i, 4).value
                labels = sheet.cell(i, 5).value
                stage_name = sheet.cell(i, 6).value
                initial_hours = sheet.cell(i, 7).value
                hours_passed = sheet.cell(i, 8).value
                hours_planned = sheet.cell(i, 9).value
                hours_remaining = sheet.cell(i, 10).value
                total_hours = sheet.cell(i, 11).value
                in_progress = sheet.cell(i, 12).value
                client = sheet.cell(i, 13).value
                client_name = sheet.cell(i, 14).value
                parent_task = sheet.cell(i, 15).value
                parent_task_title = sheet.cell(i, 16).value
                recurrent = sheet.cell(i, 17).value
                email_client = sheet.cell(i, 18).value
                start_date = sheet.cell(i, 19).value
                end_date = sheet.cell(i, 20).value
                phone = sheet.cell(i, 21).value
                sequence = sheet.cell(i, 22).value
                cc_emails = sheet.cell(i, 23).value
                visible_for = sheet.cell(i, 24).value
                visible_for_name = sheet.cell(i, 25).value
                recurrence_update = sheet.cell(i, 26).value
                task_completed = sheet.cell(i, 27).value
                time_sheets = sheet.cell(i, 28).value
                time_sheet_date = sheet.cell(i, 29).value
                time_sheet_employee_name = sheet.cell(i, 30).value
                time_sheet_description = sheet.cell(i, 31).value
                time_sheet_amount = sheet.cell(i, 32).value
                time_sheet_partner_name = sheet.cell(i, 33).value
                time_sheet_user_name = sheet.cell(i, 34).value
                work_sheet_name = sheet.cell(i, 35).value
                helpdesk_ticket_subject = sheet.cell(i, 36).value
                if end_date:
                    end_date_str = xldate_as_datetime(end_date, wb.datemode).date()
                else:
                    end_date_str = False
                if time_sheet_date:
                    time_sheet_date = xldate_as_datetime(time_sheet_date, wb.datemode).date()
                else:
                    time_sheet_date = False
                if start_date:
                    start_date_str = xldate_as_datetime(start_date, wb.datemode).date()
                else:
                    start_date_str = False
                # phone = sheet.cell(i, 21).value
                # sequence = sheet.cell(i, 22).value
                # cc_emails = sheet.cell(i, 24).value
                # visible_for = sheet.cell(i, 25).value
                # visible_for_name = sheet.cell(i, 26).value
                # recurrence_update = sheet.cell(i, 27).value
                # task_done = sheet.cell(i, 28).value
                # timesheets = sheet.cell(i, 29).value
                # timesheet_date_str = sheet.cell(i, 30).value
                # if timesheet_date_str:
                #     timesheet_date_str = xldate_as_datetime(timesheet_date_str, wb.datemode).date()
                # else:
                #     timesheet_date_str = False
                # timesheet_employee_name = sheet.cell(i, 31).value
                # timesheet_description = sheet.cell(i, 32).value
                # timesheet_amount = sheet.cell(i, 33).value
                # timesheet_partner_name = sheet.cell(i, 34).value
                # timesheet_user_name = sheet.cell(i, 35).value
                # worksheet_template_name = sheet.cell(i, 36).value
                # ticket_helpdesk_subject = sheet.cell(i, 37).value
                assigned_to_name = assigned_to

                # Condition 1: Map assigned_to
                if assigned_to == "Stephane Verbrugghe":
                    assigned_to_name = "Florence Pau"

                # Condition 2: Map client_contact_name
                if assigned_to == "Victor Lima":
                    client_contact_name = "Victor"

                tag_ids = []
                tags = []

                if company == "Pau Digital Factory":
                    tags.extend(["ERP", "WEB"])

                # Condition 5: Map tags based on company_name
                if company == 'Action Informatique':
                    tags.append("INFORMATIQUE")

                for tag in tags:
                    tag = self.env['project.tags'].search([('name', '=', tag)], limit=1)
                    tag_ids.append(tag.id)

                # Find or create related records (as before)
                project = self.env['project.project'].search([('name', '=', project)], limit=1)
                assigned_to = self.env['res.users'].search([('name', '=', assigned_to_name)], limit=1)
                company = self.env['res.company'].search([('name', '=', company)], limit=1)
                client = self.env['res.partner'].search([('name', '=', client)], limit=1)
                # client_contact = self.env['res.partner'].search([('name', '=', client_contact_name)], limit=1)
                parent_task = self.env['project.task'].search(
                    [('name', '=', parent_task_title), ('project_id', '=', project.id)], limit=1)
                stage = self.env['project.task.type'].search([('name', '=', stage_name)], limit=1)
                visible_for = self.env['res.users'].search([('name', '=', visible_for_name)], limit=1)

                # Rest of the code (as before)

                # Create or update the task
                if title:
                    # Create or update the task
                    if assigned_to:
                        assigned_to_list = [assigned_to.id]
                    else:
                        assigned_to_list = False
                    task = self.env['project.task'].search(
                        [('name', '=', title), ('project_id', '=', project.id), ('user_ids', '=', assigned_to_list)],
                        limit=1)
                    if title and task.id == 9770:
                        print(in_progress)
                    task.write({'progress': in_progress, 'planned_hours': initial_hours, 'effective_hours': hours_passed})
                    if task and parent_task and not task.parent_id:

                        task.write({'parent_id': parent_task.id})
                    print(in_progress)

                    # if not task:
                    #
                    #     task_vals = {
                    #         'name': title,
                    #         'project_id': project.id,
                    #         'user_ids': assigned_to_list,
                    #         'progress': in_progress,
                    #         'company_id': self.env.company.id,
                    #         'parent_id': parent_task.id,
                    #         'partner_id': client.id,
                    #         'stage_id': stage.id,
                    #         'partner_phone': phone,
                    #         'remaining_hours': hours_remaining,
                    #         'planned_hours': hours_planned,
                    #         'total_hours_spent': hours_passed,
                    #         'recurring_task': recurrent,
                    #         # 'total_hours_spent': hours_passed,
                    #         'tag_ids': tag_ids,
                    #         'allocated_hours': initial_hours,
                    #         'email_cc': email_client,
                    #         'date_deadline': end_date_str,
                    #         'date_assign': start_date_str,  # Adding date_start
                    #         # Add other task fields as needed
                    #     }
                    #     # try:
                    #     #     task = self.env['project.task'].create(task_vals)
                    #     #     self.env.cr.commit()
                    #     # except:
                    #     #     task = False

                # Create a timesheet line if timesheets are mentioned
                # if task and time_sheets and time_sheet_date and time_sheet_description and time_sheet_amount:
                #     existing_timesheet = self.env['account.analytic.line'].search([
                #         ('task_id', '=', task.id),
                #         ('name', '=', time_sheet_description),
                #         ('date', '=', time_sheet_date),
                #     ], limit=1)
                #     if existing_timesheet:
                #         existing_timesheet.unit_amount = time_sheet_amount
                #         self.env.cr.commit()
                #
                #     # if not existing_timesheet:
                #     #     employee_id = self.env['hr.employee'].search([('name', '=', time_sheet_employee_name)], limit=1)
                #     #     partner_id = self.env['res.partner'].search([('name', '=', time_sheet_partner_name)], limit=1)
                #     #     user_id = self.env['res.users'].search([('name', '=', time_sheet_user_name)], limit=1)
                #     #     timesheet_vals = {
                #     #         'task_id': task.id,
                #     #         'name': time_sheet_description,
                #     #         'date': time_sheet_date,
                #     #         'unit_amount': time_sheet_amount,
                #     #         'employee_id': employee_id.id,  # Add employee field if available
                #     #         'partner_id': partner_id.id,
                #     #         'user_id': user_id.id,
                #     #         # Add other fields for the timesheet as needed
                #     #     }
                #     #     try:
                #     #         timesheet = self.env['account.analytic.line'].create(timesheet_vals)
                #     #         self.env.cr.commit()
                #     #     except:
                #     #         print("timesheet_vals ", timesheet_vals)
                #     #         pass


            except Exception as e:
                # Handle exceptions or skip rows with errors
                pass

        return True
